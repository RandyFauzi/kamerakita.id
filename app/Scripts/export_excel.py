import sys
import os
import json
import openpyxl

def main():
    if len(sys.argv) < 3:
        print("Usage: python export_excel.py <json_data_path> <output_xlsx_path>")
        sys.exit(1)

    json_path = sys.argv[1]
    output_path = sys.argv[2]

    # Load JSON data
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Template path
    template_path = r"C:\laragon\www\kamerakita.id\public\Assets\Team Nanda Hourly tracker & Participant Information Indonesia.xlsx"
    if not os.path.exists(template_path):
        # Fallback to relative path in project
        template_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'public', 'Assets', 'Team Nanda Hourly tracker & Participant Information Indonesia.xlsx')

    # Load template workbook
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active # Load the active sheet 'Participats details'

    # Fill data starting from row 3
    start_row = 3
    for idx, item in enumerate(data):
        row = start_row + idx
        ws.cell(row=row, column=1, value=item.get('date_added'))
        ws.cell(row=row, column=2, value=item.get('full_name'))
        ws.cell(row=row, column=3, value=item.get('email'))
        ws.cell(row=row, column=4, value=item.get('type', 'Residential'))
        ws.cell(row=row, column=5, value=item.get('hours', 0))
        ws.cell(row=row, column=6, value=item.get('minutes', 0))
        ws.cell(row=row, column=7, value=item.get('seconds', 0))

    # Save to output path
    wb.save(output_path)
    print("SUCCESS")

if __name__ == "__main__":
    main()
